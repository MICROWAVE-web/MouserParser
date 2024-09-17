import datetime
import json
import os
import random
import time
import traceback
import warnings

warnings.filterwarnings('ignore')
import openpyxl
import pytz
import requests
from decouple import config

from proxy import getProxies, findProxy, checkProxy

xlsx_file = config('xlsx_file')
mouser_api_key = config('mouser_api_key')


def parse_stock(name, proxy):
    proxies = {
        "https": proxy,
    }
    body = json.dumps({
        "SearchByKeywordRequest": {
            "keyword": name,
            "records": 10,
            "startingRecord": 0,
            "searchOptions": "string",
            "searchWithYourSignUpLanguage": "string"
        }
    })
    headers = {"Content-Type": "application/json"}
    url = f'https://api.mouser.com/api/v1/search/keyword?apiKey={mouser_api_key}'
    r = requests.post(url, data=body, proxies=proxies, headers=headers)
    if r.status_code == 200:
        data = json.loads(r.text)
        for part in data["SearchResults"]["Parts"]:
            if part["ManufacturerPartNumber"] == name:
                if "In Stock" in part["Availability"]:
                    return int(part["Availability"].split()[0])
                else:
                    return 0
        else:
            return -1  # Обозначает не нахождение позиции


def main():
    if not os.path.exists(xlsx_file):
        raise Exception(f"Файл «{xlsx_file}» не найден.")

    proxy = None
    if not os.path.exists('proxy.txt'):
        print("Поиск нового proxy... Поиск может занять до нескольких минут.")
        proxies = getProxies()
        working_proxy = findProxy(proxies)
        with open('proxy.txt', 'w') as f:
            f.write(working_proxy)
        proxy = working_proxy
    else:
        with open('proxy.txt', 'r') as f:
            working_proxy = f.readline().strip()
        print("Проверка сохранённого proxy...")
        if checkProxy(working_proxy) is False:
            print("Поиск нового proxy... Поиск может занять до нескольких минут.")
            proxies = getProxies()
            working_proxy = findProxy(proxies)
            with open('proxy.txt', 'w') as f:
                f.write(working_proxy)
        proxy = working_proxy
    if proxy is None:
        print("PROXY не найден! (✘ᆺ✘)")
    else:
        print(f"PROXY успешно найден! (>⩊<): {working_proxy}")

    wookbook = openpyxl.load_workbook(xlsx_file)
    worksheet = wookbook.active
    for col_i in range(1, worksheet.max_column + 2):
        if worksheet.cell(row=1, column=col_i).value in ['', None]:
            n_col = col_i
            break
    worksheet.cell(row=1, column=n_col).value = datetime.datetime.now(pytz.timezone('Europe/Moscow')).strftime(
        config('time_format'))
    for row_i in range(2, worksheet.max_row + 1):
        name = worksheet.cell(row=row_i, column=2).value
        stock = parse_stock(name, proxy)
        if stock == -1:
            worksheet.cell(row=row_i, column=n_col).value = 'Ошибка'
        else:
            if bool(int(config('console_log'))):
                print(f"{name}: {stock}")
            worksheet.cell(row=row_i, column=n_col).value = stock
        time.sleep(random.randint(1, 10) // 10)
    wookbook.save(xlsx_file)
    print('Скрипт успешно завершил свою работу. ◝(ᵔᵕᵔ)◜')
    print("\nПрограмма завершит свою работу через 120 секунд.")
    time.sleep(120)


if __name__ == '__main__':
    try:
        print("Убедитесь, что в вас включен VPN. Без VPN работа скрипта невозможна.")
        main()
    except Exception:
        traceback.print_exc()
        print("\nПрограмма завершит свою работу через 120 секунд.")
        time.sleep(120)
